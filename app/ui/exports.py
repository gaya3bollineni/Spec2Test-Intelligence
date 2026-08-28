import io
import json
from copy import copy
from typing import Any

import pandas as pd
import streamlit as st
from openpyxl.styles import Font, PatternFill

from app.config import (
    EXCEL_FILE_NAME,
    EXCEL_SHEET_NAME,
    JSON_FILE_NAME,
)


TRACEABILITY_SHEET_NAME = "Traceability Matrix"


def build_export_dataframe(
    test_cases: list[Any],
    include_preconditions: bool,
) -> pd.DataFrame:
    """
    Converts generated test cases into an exportable DataFrame.
    """

    rows = []

    for test_case in test_cases:
        row = {
            "Test Case ID": test_case.test_case_id,
            "Requirement ID": test_case.requirement_id,
            "Scenario Type": test_case.scenario_type,
            "Test Scenario": test_case.test_scenario,
            "Description": test_case.test_case_description,
            "Test Steps": "\n".join(
                f"{index}. {step}"
                for index, step in enumerate(
                    test_case.test_steps,
                    start=1,
                )
            ),
            "Test Data": test_case.test_data,
            "Expected Result": test_case.expected_result,
            "Priority": test_case.priority,
        }

        if include_preconditions:
            row["Preconditions"] = "\n".join(
                f"{index}. {precondition}"
                for index, precondition in enumerate(
                    test_case.preconditions,
                    start=1,
                )
            )

        rows.append(row)

    return pd.DataFrame(rows)


def build_traceability_dataframe(
    traceability_rows: list[Any],
) -> pd.DataFrame:
    """
    Converts traceability records into an exportable DataFrame.
    """

    rows = []

    for traceability_row in traceability_rows:
        coverage_percentage = (
            traceability_row.coverage_percentage
        )

        rows.append(
            {
                "Requirement ID": traceability_row.requirement_id,
                "Acceptance Criteria": (
                    traceability_row.acceptance_criteria
                ),
                "Positive Test Cases": (
                    traceability_row.positive_count
                ),
                "Negative Test Cases": (
                    traceability_row.negative_count
                ),
                "Edge Test Cases": (
                    traceability_row.edge_count
                ),
                "Total Test Cases": (
                    traceability_row.total_test_cases
                ),
                "Coverage": (
                    f"{traceability_row.coverage_percentage}%"
                ),
                "Status": traceability_row.coverage_status,
            }
        )

    return pd.DataFrame(rows)


def format_worksheet(
    worksheet,
) -> None:
    """
    Applies basic formatting to an Excel worksheet.
    """

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )

    header_font = Font(
        bold=True,
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_cells in worksheet.columns:
        column_letter = column_cells[0].column_letter

        maximum_length = max(
            (
                len(str(cell.value))
                for cell in column_cells
                if cell.value is not None
            ),
            default=0,
        )

        worksheet.column_dimensions[column_letter].width = min(
            maximum_length + 2,
            50,
        )

    for row_cells in worksheet.iter_rows():
        for cell in row_cells:
            updated_alignment = copy(cell.alignment)
            updated_alignment.wrap_text = True
            updated_alignment.vertical = "top"
            cell.alignment = updated_alignment


def create_excel_file(
    test_case_dataframe: pd.DataFrame,
    traceability_dataframe: pd.DataFrame,
) -> io.BytesIO:
    """
    Creates an Excel workbook containing test cases and RTM sheets.
    """

    excel_buffer = io.BytesIO()

    with pd.ExcelWriter(
        excel_buffer,
        engine="openpyxl",
    ) as writer:
        test_case_dataframe.to_excel(
            writer,
            index=False,
            sheet_name=EXCEL_SHEET_NAME,
        )

        traceability_dataframe.to_excel(
            writer,
            index=False,
            sheet_name=TRACEABILITY_SHEET_NAME,
        )

        test_case_worksheet = writer.sheets[
            EXCEL_SHEET_NAME
        ]

        traceability_worksheet = writer.sheets[
            TRACEABILITY_SHEET_NAME
        ]

        format_worksheet(
            test_case_worksheet
        )

        format_worksheet(
            traceability_worksheet
        )

    excel_buffer.seek(0)

    return excel_buffer


def create_json_output(
    test_cases: list[Any],
) -> str:
    """
    Converts generated test cases into formatted JSON.
    """

    return json.dumps(
        [
            test_case.model_dump()
            for test_case in test_cases
        ],
        indent=2,
    )


def show_export_buttons(
    test_cases: list[Any],
    include_preconditions: bool,
    traceability_rows: list[Any],
) -> None:
    """
    Displays JSON and Excel download buttons.
    """

    test_case_dataframe = build_export_dataframe(
        test_cases=test_cases,
        include_preconditions=include_preconditions,
    )

    traceability_dataframe = (
        build_traceability_dataframe(
            traceability_rows=traceability_rows,
        )
    )

    json_output = create_json_output(
        test_cases
    )

    excel_output = create_excel_file(
        test_case_dataframe=test_case_dataframe,
        traceability_dataframe=traceability_dataframe,
    )

    json_column, excel_column = st.columns(2)

    with json_column:
        st.download_button(
            label="Download JSON",
            data=json_output,
            file_name=JSON_FILE_NAME,
            mime="application/json",
            width="stretch",
        )

    with excel_column:
        st.download_button(
            label="Download Excel",
            data=excel_output,
            file_name=EXCEL_FILE_NAME,
            mime=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            width="stretch",
        )